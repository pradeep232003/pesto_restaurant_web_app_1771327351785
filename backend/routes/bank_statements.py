"""
Bank Statement splitter — Manager tool.

Managers upload a bank statement (PDF / CSV / XLSX). The file is stored
in GridFS, the raw text/rows are extracted locally, and Claude Sonnet
4.5 is asked to classify each transaction as Income vs Expense, tag a
category, and (for expenses) best-match the counter-party to a known
supplier already recorded in the invoices collection.

The parsed record is persisted per location and can be downloaded as an
XLSX with **two tabs** (Income + Expenses). All endpoints are admin
gated.

Endpoints (all under /api/admin/bank-statements):
  GET    /                — list statements for a location
  POST   /upload          — upload + AI split (returns the saved record)
  GET    /{id}            — full record incl. transactions
  GET    /{id}/xlsx       — download 2-tab XLSX
  GET    /{id}/file       — stream the raw uploaded file
  DELETE /{id}            — remove statement + file
"""
import base64
import csv
import io
import json
import logging
import re
import time
import uuid
from datetime import datetime, timezone
from typing import Any, List, Optional

import gridfs
import httpx
from bson import ObjectId
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from auth import get_admin_user
from db import db
from routes import bank_statement_parser

router = APIRouter(prefix="/api/admin/bank-statements", tags=["bank-statements"])
_log = logging.getLogger("bank_statements")
_log.setLevel(logging.INFO)

_fs = gridfs.GridFS(db, collection="bank_statement_files")
statements = db["bank_statements"]
invoices_col = db["invoices"]

MAX_BYTES = 15 * 1024 * 1024  # 15 MB — bank statements are usually text-heavy

ALLOWED_CT = {
    "application/pdf",
    "text/csv",
    "application/csv",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    # some browsers send octet-stream for CSVs; we fall back on file extension
    "application/octet-stream",
    "text/plain",
}

EXPENSE_CATEGORIES = [
    "supplier", "wages", "rent", "utilities", "software",
    "repairs", "marketing", "equipment", "cleaning", "insurance",
    "bank_fees", "tax", "transfer", "other",
]

INCOME_CATEGORIES = [
    "sales", "delivery", "loyalty_topup", "refund_in",
    "grant", "transfer", "other",
]


def _strip(doc: dict) -> dict:
    out = {k: v for k, v in (doc or {}).items() if k != "_id"}
    if isinstance(out.get("file_id"), ObjectId):
        out["file_id"] = str(out["file_id"])
    return out


# ---------------------------------------------------------------------------
# File → text extraction
# ---------------------------------------------------------------------------

def _extract_text_pdf(blob: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as e:
        raise HTTPException(
            500,
            f"PDF support unavailable on this server: {e}. Ask the developer to add "
            "'pypdf' to backend/requirements-prod.txt and redeploy.",
        )
    reader = PdfReader(io.BytesIO(blob))
    pages = []
    for i, page in enumerate(reader.pages):
        # Prefer layout mode — preserves horizontal column positions
        # (crucial for UK bank statements that print separate 'Paid in'
        # and 'Paid out' columns). Fall back to plain flow if layout
        # mode chokes on the page (some scanned/malformed PDFs).
        txt = ""
        try:
            txt = page.extract_text(extraction_mode="layout") or ""
        except Exception:
            try:
                txt = page.extract_text() or ""
            except Exception:
                txt = ""
        if txt.strip():
            pages.append(f"--- Page {i + 1} ---\n{txt}")
    return "\n\n".join(pages)


def _extract_text_csv(blob: bytes) -> str:
    # Try common encodings — bank exports are messy.
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = blob.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise HTTPException(400, "Could not decode CSV file")

    # Normalise to a plain TSV-ish string for the model — cheaper tokens
    # than raw CSV with random quoting.
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return ""
    # Keep the first 800 rows only — anything more is very unusual for a
    # month's statement and blows the model context budget.
    rows = rows[:800]
    return "\n".join(["\t".join(str(c).strip() for c in r) for r in rows])


def _extract_text_xlsx(blob: bytes) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        parts.append(f"--- Sheet: {sheet.title} ---")
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i > 800:
                parts.append("... (rows truncated)")
                break
            if any(c not in (None, "") for c in row):
                parts.append("\t".join("" if c is None else str(c).strip() for c in row))
    return "\n".join(parts)


def _extract_text(blob: bytes, content_type: str, filename: str) -> str:
    ct = (content_type or "").lower()
    name = (filename or "").lower()
    if ct == "application/pdf" or name.endswith(".pdf"):
        return _extract_text_pdf(blob)
    if name.endswith(".csv") or ct in ("text/csv", "application/csv", "text/plain"):
        return _extract_text_csv(blob)
    if name.endswith(".xlsx") or name.endswith(".xls") or "spreadsheet" in ct or "excel" in ct:
        return _extract_text_xlsx(blob)
    # Last-ditch: try PDF then CSV.
    try:
        return _extract_text_pdf(blob)
    except Exception:
        return _extract_text_csv(blob)


# ---------------------------------------------------------------------------
# AI classifier — Claude Sonnet 4.5 via raw httpx (Railway-safe)
# ---------------------------------------------------------------------------

_AI_SYSTEM = (
    "You are a UK bank-statement analyst for Jolly's Kafe, a hospitality "
    "business. You receive raw text extracted from a bank statement (PDF, "
    "CSV or XLSX). Your job is to identify EVERY financial transaction and "
    "return a strict JSON split of income vs expense.\n\n"
    "===== HOW TO READ UK BANK STATEMENTS — READ CAREFULLY =====\n"
    "UK bank statements ALWAYS use two dedicated money columns:\n"
    "  • 'Paid in' / 'Money in' / 'Credit' / 'CR'  = INCOME  (type='income')\n"
    "  • 'Paid out' / 'Money out' / 'Debit' / 'DR' = EXPENSE (type='expense')\n"
    "The column POSITION on the row (or the sign/prefix) is the AUTHORITATIVE "
    "signal — NEVER override it based on the merchant name.\n"
    "\nPDF text extraction often flattens tables into space-separated text. "
    "Before classifying, locate the header row (look for the words 'Paid in', "
    "'Paid out', 'Money in', 'Money out', 'Credit', 'Debit', 'Balance') and "
    "note the horizontal position of each money column. Then for every "
    "transaction row, the amount that sits under the 'Paid in' column is "
    "INCOME and the amount under 'Paid out' is an EXPENSE — even if the "
    "description looks like a supplier name (a refund from Bidfood is still "
    "income; wages arriving from HMRC as a grant are still income).\n"
    "\nRow patterns you will see after PDF extraction:\n"
    "  DATE  DESC  <blank>  87.50  4,230.10   → 87.50 is 'Paid out' → EXPENSE\n"
    "  DATE  DESC  1,240.00  <blank>  5,470.10 → 1,240.00 is 'Paid in' → INCOME\n"
    "Some banks (Barclays, NatWest, Lloyds, Santander, HSBC) print CR / DR "
    "after the amount instead of using two columns — 'CR' = income, 'DR' = "
    "expense. Amounts with a leading minus '-' or in brackets '(…)' are "
    "expenses. A leading '+' is income.\n"
    "\nIf, after applying these rules, you still can't tell the direction of a "
    "specific row, default to expense and add a note in the description "
    "'(direction unclear)'. NEVER guess based only on the merchant.\n"
    "===== END OF DIRECTION RULES =====\n\n"
    "RULES:\n"
    "1. Output STRICT JSON, no markdown, no commentary.\n"
    "2. One row per transaction — never combine credits and debits into one "
    "row.\n"
    "3. Ignore running-balance rows and opening/closing balance summaries. "
    "The Balance column is NEVER a transaction.\n"
    "4. Dates: return ISO YYYY-MM-DD. If year is missing, infer from the "
    "statement period; if impossible, use empty string ''.\n"
    "5. Amounts are positive floats with 2 decimals — the row's `type` field "
    "carries the direction ('income' or 'expense'). Strip currency symbols, "
    "thousands separators, +/- signs and CR/DR suffixes before writing the "
    "number.\n"
    "6. `description` = the counterparty / merchant / reference line as "
    "printed. Trim whitespace but do NOT paraphrase.\n"
    "7. Best-guess a category slug (see lists below). If uncertain use "
    "'other'.\n"
    "8. For EXPENSES only: if the description clearly matches one of the "
    "KNOWN_SUPPLIERS list provided in the user message, copy that supplier "
    "name (verbatim) into `matched_supplier`. Otherwise return "
    "matched_supplier: ''.\n"
    "9. Currency defaults to GBP. Detect the statement's currency if visible "
    "and put it in the top-level `currency` field.\n\n"
    "EXPENSE CATEGORIES (slugs): supplier, wages, rent, utilities, software, "
    "repairs, marketing, equipment, cleaning, insurance, bank_fees, tax, "
    "transfer, other.\n"
    "INCOME CATEGORIES (slugs): sales, delivery, loyalty_topup, refund_in, "
    "grant, transfer, other.\n\n"
    "RESPONSE SHAPE:\n"
    "{\n"
    "  \"period_start\": \"YYYY-MM-DD or ''\",\n"
    "  \"period_end\":   \"YYYY-MM-DD or ''\",\n"
    "  \"account_ref\":  \"account name / number if visible, else ''\",\n"
    "  \"currency\":     \"GBP\",\n"
    "  \"transactions\": [\n"
    "    {\n"
    "      \"date\": \"YYYY-MM-DD\",\n"
    "      \"description\": \"AMAZON UK MARKETPL\",\n"
    "      \"amount\": 24.99,\n"
    "      \"type\": \"expense\",\n"
    "      \"category\": \"supplier\",\n"
    "      \"matched_supplier\": \"\"\n"
    "    }\n"
    "  ]\n"
    "}\n"
)


def _scrub_json(text: str) -> str:
    """Peel markdown fences off Claude's response so json.loads works."""
    text = text.strip()
    if text.startswith("```"):
        # ```json\n...\n``` or ```\n...\n```
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    return text.strip()


def _known_suppliers(location_id: str) -> List[str]:
    """Distinct supplier names seen on invoices for this location."""
    try:
        names = invoices_col.distinct("supplier", {"location_id": location_id})
    except Exception:
        return []
    return sorted({(n or "").strip() for n in names if n and str(n).strip()})[:200]


# Common corporate suffixes / tokens to strip when comparing supplier
# names against transaction descriptions. Keeps things simple — we don't
# need Levenshtein, just uppercase + suffix strip + substring match.
_SUPPLIER_STOPWORDS = {
    "LTD", "LIMITED", "PLC", "LLC", "LLP", "INC", "INCORPORATED",
    "CO", "COMPANY", "UK", "GB", "AND", "&", "THE",
}


def _normalise_supplier(name: str) -> str:
    """Uppercase, strip punctuation and common corporate suffixes so
    'BOOKER LTD' and 'Booker Wholesale Ltd' both collapse to 'BOOKER'."""
    if not name:
        return ""
    s = name.upper()
    s = re.sub(r"[.,'\"()\[\]{}/\\\-_&]+", " ", s)
    tokens = [t for t in s.split() if t and t not in _SUPPLIER_STOPWORDS]
    return " ".join(tokens)


def _match_supplier(description: str, suppliers: List[str]) -> str:
    """Best-effort fuzzy match: normalise both sides, then look for a
    whole-word supplier fragment inside the description. The longest
    matching supplier wins (so 'Booker Wholesale' beats plain 'Booker'
    when both are known). Returns the ORIGINAL (unnormalised) supplier
    name so what the user sees on invoices stays consistent.
    """
    norm_desc = _normalise_supplier(description)
    if not norm_desc:
        return ""
    best: tuple = ("", 0)  # (original_name, score = length of match)
    desc_tokens = set(norm_desc.split())
    for orig in suppliers:
        norm_sup = _normalise_supplier(orig)
        if not norm_sup:
            continue
        sup_tokens = norm_sup.split()
        # Require ALL supplier tokens to appear as whole words in the
        # description (order-independent). Cheap but effective — avoids
        # false hits like matching "Booker" against "Bookworm Cafe".
        if all(t in desc_tokens for t in sup_tokens):
            score = len(norm_sup)
            if score > best[1]:
                best = (orig, score)
    return best[0]


def _load_invoices(location_id: str) -> List[dict]:
    """Fetch scanned + uploaded invoices for a location — minimal
    projection so we don't pull GridFS blobs. Used to match bank-statement
    expense rows to actual invoices."""
    try:
        rows = list(invoices_col.find(
            {"location_id": location_id},
            {
                "id": 1, "supplier": 1, "invoice_number": 1,
                "invoice_date": 1, "total": 1, "grand_total": 1,
            },
        ).limit(2000))
    except Exception:
        return []
    out = []
    for r in rows:
        supplier = (r.get("supplier") or "").strip()
        if not supplier:
            continue
        # invoice value — most records use `total`; some legacy use `grand_total`.
        amt = r.get("total") or r.get("grand_total") or 0
        try:
            amt = float(amt)
        except (TypeError, ValueError):
            amt = 0.0
        out.append({
            "id": r.get("id", ""),
            "supplier": supplier,
            "invoice_number": (r.get("invoice_number") or "").strip(),
            "invoice_date": (r.get("invoice_date") or "").strip(),
            "amount": round(amt, 2),
        })
    return out


def _match_invoice(description: str, txn_amount: float, txn_date: str,
                   invoices: List[dict]) -> Optional[dict]:
    """Find the single best invoice match for a bank-statement expense.

    Strategy (all fuzzy):
      1. Normalise description AND supplier names, then find every
         invoice whose supplier tokens all appear in the description.
      2. Among those, prefer the invoice whose amount matches to within
         ±£0.05, then whose date is closest to the txn date.
      3. If no candidates by supplier, fall back to amount+date only
         within a tighter tolerance (rare — for descriptions where the
         supplier text is missing entirely).
    Returns the matching invoice dict, or None.
    """
    if not invoices:
        return None

    def _parse_date(s: str) -> Optional[datetime]:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s[:10], fmt)
            except Exception:
                continue
        return None

    txn_dt = _parse_date(txn_date) if txn_date else None
    norm_desc = _normalise_supplier(description)
    desc_tokens = set(norm_desc.split())

    # Pass 1: supplier-token match
    supplier_candidates = []
    for inv in invoices:
        norm_sup = _normalise_supplier(inv["supplier"])
        if not norm_sup:
            continue
        sup_tokens = norm_sup.split()
        if all(t in desc_tokens for t in sup_tokens):
            supplier_candidates.append(inv)

    def _rank(candidate: dict) -> tuple:
        amt_diff = abs(candidate["amount"] - txn_amount)
        # Prefer exact amount matches (within 5p)
        amt_score = 0 if amt_diff < 0.05 else amt_diff
        # Date proximity in days (large penalty if we can't parse)
        inv_dt = _parse_date(candidate["invoice_date"])
        if inv_dt and txn_dt:
            date_score = abs((inv_dt - txn_dt).days)
        else:
            date_score = 999
        # Also prefer longer supplier names (more specific match)
        specificity = -len(_normalise_supplier(candidate["supplier"]))
        return (amt_score, date_score, specificity)

    if supplier_candidates:
        return sorted(supplier_candidates, key=_rank)[0]

    # Pass 2: amount + date fallback (rare)
    if txn_dt and txn_amount > 0:
        loose = []
        for inv in invoices:
            if abs(inv["amount"] - txn_amount) > 0.05:
                continue
            inv_dt = _parse_date(inv["invoice_date"])
            if not inv_dt:
                continue
            if abs((inv_dt - txn_dt).days) <= 7:
                loose.append(inv)
        if loose:
            return sorted(loose, key=_rank)[0]
    return None


def _format_invoice_ref(inv: dict) -> str:
    """Pretty display string for a matched invoice — 'Bidfood #INV1234'."""
    if not inv:
        return ""
    supplier = inv.get("supplier") or ""
    number = inv.get("invoice_number") or ""
    if supplier and number:
        return f"{supplier} #{number}"
    return supplier or number or ""


def _enrich_with_invoices(txns: List[dict], location_id: str) -> int:
    """Match every EXPENSE txn against real invoice records for the
    location. Updates each txn in place with `matched_invoice_id`,
    `matched_invoice_ref` ("Supplier #INVNUM") and keeps
    `matched_supplier` in sync (used by the Supplier pivot).

    Returns the number of matches made — logged upstream so managers
    can see how many statement rows the system found paperwork for.

    IMPORTANT: any stale invoice reference stored on the txn (from a
    previous classification, where the invoice may since have been
    deleted or edited) is cleared BEFORE re-matching, so re-classify
    always reflects the current state of the invoices collection.
    """
    invoices = _load_invoices(location_id)
    valid_ids = {inv.get("id") for inv in invoices if inv.get("id")}
    matches = 0
    for t in txns:
        if t.get("type") != "expense":
            continue
        # Drop any stale invoice reference that points at a deleted
        # invoice, so the "Matched Invoice" column never shows an
        # orphaned reference.
        stale_id = t.get("matched_invoice_id")
        if stale_id and stale_id not in valid_ids:
            t.pop("matched_invoice_id", None)
            t.pop("matched_invoice_ref", None)
        if not invoices:
            continue
        inv = _match_invoice(
            description=t.get("description", ""),
            txn_amount=float(t.get("amount") or 0),
            txn_date=t.get("date", ""),
            invoices=invoices,
        )
        if not inv:
            # No new match — leave any pre-existing supplier fuzzy-match
            # alone but ensure the invoice-specific fields are clean.
            t.pop("matched_invoice_id", None)
            t.pop("matched_invoice_ref", None)
            continue
        t["matched_invoice_id"] = inv.get("id", "")
        t["matched_invoice_ref"] = _format_invoice_ref(inv)
        if not t.get("matched_supplier"):
            t["matched_supplier"] = inv.get("supplier", "")
        matches += 1
    return matches


def _chunk_text(text: str, target_size: int = 15_000) -> List[str]:
    """Split extracted statement text into ~15 kB chunks along natural
    page boundaries (`--- Page N ---` markers). Falls back to line-based
    chunking if no page markers are present (CSV/XLSX). Each chunk keeps
    the AI response well under the 8 000-token cap, and parallel calls
    stay comfortably under Cloudflare's 100 s ingress timeout.
    """
    if len(text) <= target_size:
        return [text]

    # Prefer page-marker splits — cleanest for PDFs.
    page_split = re.split(r"(?=--- Page \d+)", text)
    page_split = [p for p in page_split if p.strip()]

    # Fallback: split by lines if no markers detected.
    if len(page_split) <= 1:
        lines = text.splitlines()
        page_split = []
        cur: list = []
        cur_len = 0
        for ln in lines:
            cur.append(ln)
            cur_len += len(ln) + 1
            if cur_len >= target_size:
                page_split.append("\n".join(cur))
                cur, cur_len = [], 0
        if cur:
            page_split.append("\n".join(cur))

    # Now pack consecutive small pages together up to target_size.
    chunks: List[str] = []
    buf: list = []
    buf_len = 0
    for piece in page_split:
        piece_len = len(piece)
        if buf and buf_len + piece_len > target_size:
            chunks.append("".join(buf) if not buf[0].startswith("--- Page") else "\n".join(buf))
            buf, buf_len = [], 0
        buf.append(piece)
        buf_len += piece_len
    if buf:
        chunks.append("".join(buf) if not buf[0].startswith("--- Page") else "\n".join(buf))
    return [c.strip() for c in chunks if c.strip()]


async def _classify_chunk(chunk_text: str, chunk_idx: int, total_chunks: int,
                          supplier_hint: str, api_key: str, provider: str,
                          depth: int = 0) -> dict:
    """Send a single chunk to Claude and return the parsed transactions
    (+ optional meta fields). Isolated in its own coroutine so multiple
    chunks can run in parallel via asyncio.gather().

    If Claude truncates the response at the max_tokens boundary (very
    dense chunks), the chunk is split in half and each half retried
    recursively (max depth 3). Results are merged transparently so the
    caller sees a single dict.
    """
    part_note = (
        f"NOTE: This is chunk {chunk_idx} of {total_chunks} from a larger bank "
        f"statement. Return transactions ONLY from THIS chunk. If period_start "
        f"/ period_end / account_ref / currency are visible in this chunk, "
        f"include them; otherwise return empty strings.\n\n"
        if total_chunks > 1 else ""
    )
    user_msg = (
        supplier_hint
        + part_note
        + "BANK STATEMENT TEXT:\n"
        + chunk_text
        + "\n\nReturn strict JSON per the system schema. No prose."
    )
    req = {
        "model": "claude-sonnet-4-5-20250929",
        "max_tokens": 8000,
        "system": _AI_SYSTEM,
        "messages": [{"role": "user", "content": user_msg}],
    }
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    async with httpx.AsyncClient(timeout=120.0) as client:
        resp = await client.post("https://api.anthropic.com/v1/messages", json=req, headers=headers)

    if resp.status_code >= 400:
        try:
            err = resp.json().get("error", {}).get("message", resp.text)
        except Exception:
            err = resp.text
        raise HTTPException(500, f"AI provider error ({provider}, {resp.status_code}) on chunk {chunk_idx}: {err[:240]}")

    data = resp.json()
    stop_reason = data.get("stop_reason")
    out = "".join(
        block.get("text", "") for block in (data.get("content") or [])
        if block.get("type") == "text"
    ).strip()
    if not out:
        raise HTTPException(500, f"AI returned an empty response for chunk {chunk_idx}")

    # If Claude hit the output-token cap, JSON is guaranteed to be
    # unterminated. Split the chunk in half and retry each half —
    # recursion depth capped at 3 (i.e. up to 8× subdivision) which is
    # more than enough for the densest real-world statements.
    truncated = stop_reason == "max_tokens"
    parsed = None
    if not truncated:
        try:
            parsed = json.loads(_scrub_json(out))
        except json.JSONDecodeError:
            # Model produced non-JSON despite `stop_reason=end_turn`.
            # Almost always still a hidden truncation; treat as such
            # and split, unless we're already very small.
            truncated = True

    if truncated:
        if depth >= 3 or len(chunk_text) < 1500:
            snippet = out[-160:].replace("\n", " ")
            raise HTTPException(
                500,
                f"Chunk {chunk_idx} truncated by AI even after {depth} sub-splits "
                f"(chunk size {len(chunk_text)} chars). Last output fragment: {snippet!r}",
            )
        _log.info("classify: chunk %d truncated (depth %d), sub-splitting %d chars",
                  chunk_idx, depth, len(chunk_text))
        half = len(chunk_text) // 2
        # Prefer to break at a newline so page/row context is preserved.
        break_at = chunk_text.rfind("\n", int(half * 0.6), int(half * 1.4))
        if break_at < 0:
            break_at = half
        left, right = chunk_text[:break_at], chunk_text[break_at:]
        a = await _classify_chunk(left, chunk_idx, total_chunks, supplier_hint, api_key, provider, depth + 1)
        b = await _classify_chunk(right, chunk_idx, total_chunks, supplier_hint, api_key, provider, depth + 1)
        merged: dict = {"period_start": "", "period_end": "", "account_ref": "", "currency": "", "transactions": []}
        for p in (a, b):
            for k in ("period_start", "period_end", "account_ref", "currency"):
                if not merged[k] and (p.get(k) or "").strip():
                    merged[k] = (p.get(k) or "").strip()
            merged["transactions"].extend(p.get("transactions") or [])
        return merged

    return parsed


async def _classify(text: str, location_id: str) -> dict:
    import asyncio
    from routes.ai_settings import get_active_ai_key, get_active_ai_provider
    api_key = get_active_ai_key()
    if not api_key:
        raise HTTPException(
            500,
            "AI bank-statement splitter unavailable: no API key configured. "
            "Open Admin → AI Settings to add one.",
        )
    provider = get_active_ai_provider()

    if not text or not text.strip():
        raise HTTPException(400, "Could not extract any text from the file")

    # Hard-cap total input at ~250 k chars (~60k tokens) — covers a full
    # year of transactions. Anything beyond is very unusual and would
    # bust the 200 k context anyway.
    if len(text) > 250_000:
        text = text[:250_000] + "\n\n... (truncated by server)"

    suppliers = _known_suppliers(location_id)
    supplier_hint = ""
    if suppliers:
        supplier_hint = (
            "KNOWN_SUPPLIERS (use exact match if the description clearly "
            "references one of these — expense rows only):\n- "
            + "\n- ".join(suppliers)
            + "\n\n"
        )

    chunks = _chunk_text(text, target_size=9_000)
    total = len(chunks)
    _log.info(
        "classify: text=%d chars, suppliers=%d, chunks=%d",
        len(text), len(suppliers), total,
    )

    # Bounded concurrency — 4 chunks in parallel keeps us under
    # Anthropic's per-tier rate limits while still finishing a 20-page
    # statement in ~30-40 s wall-clock.
    sem = asyncio.Semaphore(4)

    async def _one(idx_text):
        idx, chunk = idx_text
        async with sem:
            t0 = time.monotonic()
            try:
                res = await _classify_chunk(chunk, idx, total, supplier_hint, api_key, provider)
                dur = time.monotonic() - t0
                n = len(res.get("transactions") or []) if isinstance(res, dict) else 0
                _log.info("classify: chunk %d/%d OK in %.1fs · %d txns", idx, total, dur, n)
                return res
            except HTTPException as he:
                dur = time.monotonic() - t0
                _log.warning("classify: chunk %d/%d FAILED in %.1fs · %s", idx, total, dur, he.detail)
                raise
            except Exception as e:  # noqa: BLE001
                dur = time.monotonic() - t0
                snippet = (str(e) or e.__class__.__name__).splitlines()[0][:240]
                _log.exception("classify: chunk %d/%d CRASHED in %.1fs · %s", idx, total, dur, snippet)
                raise HTTPException(500, f"AI provider error ({provider}) on chunk {idx}: {snippet}")

    t_start = time.monotonic()
    try:
        parts = await asyncio.gather(*[_one((i + 1, c)) for i, c in enumerate(chunks)])
    except HTTPException:
        raise
    _log.info("classify: all %d chunks done in %.1fs", total, time.monotonic() - t_start)

    # Merge — first non-empty meta wins; concatenate transactions.
    merged: dict = {
        "period_start": "",
        "period_end": "",
        "account_ref": "",
        "currency": "GBP",
        "transactions": [],
    }
    for p in parts:
        if not isinstance(p, dict):
            continue
        for k in ("period_start", "period_end", "account_ref", "currency"):
            if not merged[k] and (p.get(k) or "").strip():
                merged[k] = (p.get(k) or "").strip()
        txns = p.get("transactions") or []
        if isinstance(txns, list):
            merged["transactions"].extend(txns)

    txns_raw = merged["transactions"]
    if not isinstance(txns_raw, list):
        raise HTTPException(500, "AI response missing 'transactions' array")

    txns: List[dict] = []
    for t in txns_raw:
        if not isinstance(t, dict):
            continue
        desc = (t.get("description") or "").strip()
        if not desc:
            continue
        try:
            amt = round(abs(float(t.get("amount") or 0)), 2)
        except (TypeError, ValueError):
            amt = 0.0
        if amt <= 0:
            continue
        ttype = (t.get("type") or "").strip().lower()
        if ttype not in ("income", "expense"):
            # Default to expense — safer for hospitality where most rows
            # are outgoings; manager can flip if needed.
            ttype = "expense"
        cat = (t.get("category") or "").strip().lower() or "other"
        if ttype == "expense" and cat not in EXPENSE_CATEGORIES:
            cat = "other"
        if ttype == "income" and cat not in INCOME_CATEGORIES:
            cat = "other"
        supplier = (t.get("matched_supplier") or "").strip() if ttype == "expense" else ""
        # If the AI didn't tag a supplier for this expense, run our own
        # normalised fuzzy match (uppercase, strip 'LTD'/'LIMITED'/'PLC'
        # etc.) so 'BOOKER LTD' still maps to a known 'Booker' invoice
        # supplier even when the AI plays it safe.
        if ttype == "expense" and not supplier and suppliers:
            supplier = _match_supplier(desc, suppliers)
        txns.append({
            "date": (t.get("date") or "").strip(),
            "description": desc,
            "amount": amt,
            "type": ttype,
            "category": cat,
            "matched_supplier": supplier,
        })

    # Deduplicate — parallel chunks with overlapping page splits could
    # theoretically emit the same row twice. Cheap fingerprint on
    # (date, description, amount, type).
    seen: set = set()
    deduped: List[dict] = []
    for t in txns:
        key = (t["date"], t["description"], t["amount"], t["type"])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(t)

    return {
        "period_start": merged["period_start"],
        "period_end": merged["period_end"],
        "account_ref": merged["account_ref"],
        "currency": merged["currency"] or "GBP",
        "transactions": deduped,
    }


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Custom rules — global (one rule set for the whole business per user's
# earlier choice). Kept in the `bank_statement_rules` Mongo collection.
# Registered BEFORE the /{sid}/* paths so 'rules' never gets swallowed as
# a statement id.
# ---------------------------------------------------------------------------
rules_col = db["bank_statement_rules"]

_ALLOWED_EXPENSE_CATS = [
    "supplier", "wages", "rent", "utilities", "software",
    "repairs", "marketing", "equipment", "cleaning", "insurance",
    "bank_fees", "tax", "transfer", "other",
]
_ALLOWED_INCOME_CATS = [
    "sales", "delivery", "loyalty_topup", "refund_in", "grant", "transfer", "other",
]


def _load_custom_rules() -> List[dict]:
    """Fetch and compile all active custom rules once per request. Rules
    that fail to compile (bad regex) are silently skipped — the UI shows
    a compile-error flag on the row so the manager can fix it."""
    out: List[dict] = []
    for r in rules_col.find({"disabled": {"$ne": True}}).sort("created_at", 1):
        try:
            compiled = bank_statement_parser.compile_custom_rule(
                r.get("pattern", ""), r.get("mode", "simple"),
            )
        except Exception:
            continue
        out.append({
            "label": r.get("label", ""),
            "type": r.get("type", "expense"),
            "category": r.get("category", "other"),
            "_compiled": compiled,
        })
    return out


@router.get("/rules")
async def list_rules(user: dict = Depends(get_admin_user)):
    """Return built-in + custom rules for display in the Manager UI."""
    builtins = bank_statement_parser.get_builtin_rules()
    custom_docs = list(rules_col.find().sort("created_at", -1))
    customs: List[dict] = []
    for r in custom_docs:
        # Test-compile so we can surface a broken-regex badge in the UI.
        compile_error = ""
        try:
            bank_statement_parser.compile_custom_rule(
                r.get("pattern", ""), r.get("mode", "simple"),
            )
        except Exception as e:  # noqa: BLE001
            compile_error = str(e)[:180]
        customs.append({
            "id": r.get("id"),
            "label": r.get("label", ""),
            "pattern": r.get("pattern", ""),
            "mode": r.get("mode", "simple"),
            "type": r.get("type", "expense"),
            "category": r.get("category", "other"),
            "disabled": bool(r.get("disabled")),
            "created_at": r.get("created_at", ""),
            "created_by_name": r.get("created_by_name", ""),
            "compile_error": compile_error,
            "builtin": False,
        })
    return {
        "custom": customs,
        "builtin": builtins,
        "expense_categories": _ALLOWED_EXPENSE_CATS,
        "income_categories": _ALLOWED_INCOME_CATS,
    }


class RuleIn(BaseModel):
    label: str
    pattern: str
    type: str          # 'income' | 'expense'
    category: str      # slug from _ALLOWED_*_CATS
    mode: Optional[str] = "simple"  # 'simple' | 'regex'


@router.post("/rules")
async def create_rule(body: RuleIn, user: dict = Depends(get_admin_user)):
    label = (body.label or "").strip()[:80]
    pattern = (body.pattern or "").strip()[:400]
    ttype = (body.type or "").strip().lower()
    cat = (body.category or "").strip().lower()
    mode = (body.mode or "simple").strip().lower()
    if mode not in ("simple", "regex"):
        mode = "simple"
    if ttype not in ("income", "expense"):
        raise HTTPException(400, "type must be 'income' or 'expense'")
    allowed = _ALLOWED_INCOME_CATS if ttype == "income" else _ALLOWED_EXPENSE_CATS
    if cat not in allowed:
        raise HTTPException(400, f"category must be one of {allowed}")
    if not label:
        raise HTTPException(400, "label is required")
    if not pattern:
        raise HTTPException(400, "pattern is required")

    # Validate the pattern compiles before saving.
    try:
        bank_statement_parser.compile_custom_rule(pattern, mode)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(400, f"Invalid pattern: {str(e)[:160]}")

    doc = {
        "id": str(uuid.uuid4())[:12],
        "label": label,
        "pattern": pattern,
        "mode": mode,
        "type": ttype,
        "category": cat,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user.get("email", ""),
        "created_by_name": user.get("name", ""),
    }
    rules_col.insert_one(dict(doc))
    _log.info("rule created: label=%r type=%s cat=%s by=%s", label, ttype, cat, user.get("email"))
    return doc


@router.delete("/rules/{rid}")
async def delete_rule(rid: str, user: dict = Depends(get_admin_user)):
    res = rules_col.delete_one({"id": rid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Rule not found")
    _log.info("rule deleted: id=%s by=%s", rid, user.get("email"))
    return {"deleted": True}


@router.get("")
async def list_statements(
    location_id: Optional[str] = None,
    location_ids: Optional[str] = None,  # comma-separated list — multi-select from Details tab
    user: dict = Depends(get_admin_user),
):
    q: dict = {}
    ids: List[str] = []
    if location_ids:
        ids = [x.strip() for x in location_ids.split(",") if x.strip()]
    if location_id:
        ids.append(location_id)
    ids = list({i for i in ids if i})
    if ids:
        q["location_id"] = {"$in": ids} if len(ids) > 1 else ids[0]
    rows = list(statements.find(q, {"transactions": 0}).sort("uploaded_at", -1).limit(200))
    return {"items": [_strip(r) for r in rows]}


@router.get("/aggregate")
async def aggregate_statements(
    location_ids: Optional[str] = None,
    ids: Optional[str] = None,  # comma-separated statement ids — takes precedence
    user: dict = Depends(get_admin_user),
):
    """Combined transaction feed across multiple statements. Used by the
    Details tab: when the manager filters by one or many sites (and/or
    picks specific statements) the frontend gets one merged transaction
    stream with the source statement id/filename tagged on each row.
    """
    stmt_ids = [x.strip() for x in (ids or "").split(",") if x.strip()]
    loc_ids = [x.strip() for x in (location_ids or "").split(",") if x.strip()]

    q: dict = {}
    if stmt_ids:
        q["id"] = {"$in": stmt_ids}
    elif loc_ids:
        q["location_id"] = {"$in": loc_ids} if len(loc_ids) > 1 else loc_ids[0]

    rows = list(statements.find(q).sort("uploaded_at", -1))
    merged: list = []
    for r in rows:
        for t in (r.get("transactions") or []):
            merged.append({
                **t,
                "statement_id": r.get("id"),
                "statement_filename": r.get("filename"),
                "location_id": r.get("location_id"),
            })

    inc = [t for t in merged if t.get("type") == "income"]
    exp = [t for t in merged if t.get("type") == "expense"]
    return {
        "statement_count": len(rows),
        "transactions": merged,
        "income_count": len(inc),
        "expense_count": len(exp),
        "total_income": round(sum(float(t.get("amount") or 0) for t in inc), 2),
        "total_expense": round(sum(float(t.get("amount") or 0) for t in exp), 2),
    }


@router.get("/aggregate/xlsx")
async def aggregate_xlsx(
    location_ids: Optional[str] = None,
    ids: Optional[str] = None,
    user: dict = Depends(get_admin_user),
):
    """Combined XLSX across many statements — same 3-tab format used by
    the per-statement download, but with Statement + Site columns
    appended on the Income and Expenses tabs so it's obvious which row
    came from where. Query params mirror `/aggregate`.

    NOTE: registered BEFORE the `/{sid}/*` routes so FastAPI matches this
    static path first (otherwise `sid=aggregate` would swallow it).
    """
    stmt_ids = [x.strip() for x in (ids or "").split(",") if x.strip()]
    loc_ids = [x.strip() for x in (location_ids or "").split(",") if x.strip()]

    q: dict = {}
    if stmt_ids:
        q["id"] = {"$in": stmt_ids}
    elif loc_ids:
        q["location_id"] = {"$in": loc_ids} if len(loc_ids) > 1 else loc_ids[0]

    rows = list(statements.find(q).sort("uploaded_at", -1))
    if not rows:
        raise HTTPException(404, "No statements match those filters")

    income_txns: list = []
    expense_txns: list = []
    for r in rows:
        for t in (r.get("transactions") or []):
            tagged = {
                **t,
                "statement_id": r.get("id"),
                "statement_filename": r.get("filename"),
                "location_id": r.get("location_id"),
            }
            if t.get("type") == "income":
                income_txns.append(tagged)
            elif t.get("type") == "expense":
                expense_txns.append(tagged)

    total_income = round(sum(float(t.get("amount") or 0) for t in income_txns), 2)
    total_expense = round(sum(float(t.get("amount") or 0) for t in expense_txns), 2)
    locs_used = sorted({r.get("location_id", "") for r in rows if r.get("location_id")})

    summary_rows = [
        ("Statements included", len(rows)),
        ("Sites", ", ".join(locs_used) or "—"),
        ("Income transactions", len(income_txns)),
        ("Expense transactions", len(expense_txns)),
        ("Total income", total_income),
        ("Total expense", total_expense),
        ("Net", round(total_income - total_expense, 2)),
        ("Generated at", datetime.now(timezone.utc).isoformat()),
        ("Generated by", user.get("name") or user.get("email", "")),
    ]
    try:
        buf = _build_split_xlsx(income_txns, expense_txns, summary_rows)
    except Exception as ex:  # pragma: no cover — surfaced to caller
        _log.exception("Aggregate XLSX build failed for %d statements", len(rows))
        raise HTTPException(500, f"XLSX build failed: {type(ex).__name__}: {ex}") from ex

    stem = "combined" if len(rows) > 1 else re.sub(
        r"[^A-Za-z0-9._-]+", "_",
        (rows[0].get("filename") or "statement").rsplit(".", 1)[0],
    )[:60]
    date_tag = datetime.now(timezone.utc).strftime("%Y%m%d")
    out_name = f"{stem}_{len(rows)}stmts_{date_tag}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{out_name}"'},
    )


@router.post("/upload")
async def upload_statement(
    file: UploadFile = File(...),
    location_id: str = Form(...),
    engine: str = Form("python"),  # 'python' (default, free) | 'ai' (Claude fallback)
    user: dict = Depends(get_admin_user),
):
    if not location_id:
        raise HTTPException(400, "location_id is required")

    upload_t0 = time.monotonic()
    blob = await file.read()
    if not blob:
        raise HTTPException(400, "Empty file")
    if len(blob) > MAX_BYTES:
        raise HTTPException(400, f"File too large (max {MAX_BYTES // (1024 * 1024)} MB)")

    ct = file.content_type or "application/octet-stream"
    fname = file.filename or "statement"
    engine = (engine or "python").lower()
    if engine not in ("python", "ai"):
        engine = "python"
    _log.info(
        "upload: user=%s file=%r size=%d ct=%s loc=%s engine=%s",
        user.get("email", "?"), fname, len(blob), ct, location_id, engine,
    )
    if ct not in ALLOWED_CT and not any(fname.lower().endswith(x) for x in (".pdf", ".csv", ".xlsx", ".xls")):
        raise HTTPException(400, f"Unsupported file type: {ct}. Upload PDF, CSV or XLSX.")

    # Persist the file in GridFS FIRST so the manager can re-download the
    # original later — and so the AI fallback (via /{sid}/reclassify-ai)
    # has something to work with even if the Python parser succeeded.
    file_id = _fs.put(blob, filename=fname, content_type=ct)

    strategy = ""
    parsed: dict
    if engine == "python":
        # Free deterministic parser — no AI tokens burnt.
        try:
            t_parse = time.monotonic()
            suppliers = _known_suppliers(location_id)
            custom_rules = _load_custom_rules()
            parsed = bank_statement_parser.parse_locally(
                blob=blob, content_type=ct, filename=fname,
                suppliers=suppliers, supplier_matcher=_match_supplier,
                custom_rules=custom_rules,
            )
            strategy = parsed.pop("_strategy", "python")
            n_matched = _enrich_with_invoices(parsed["transactions"], location_id)
            _log.info("upload: python parser %s → %d txns in %.1fs (%d custom rules, %d invoice matches)",
                      strategy, len(parsed["transactions"]), time.monotonic() - t_parse,
                      len(custom_rules), n_matched)
        except ValueError as ve:
            try:
                _fs.delete(file_id)
            except Exception:
                pass
            raise HTTPException(422, str(ve))
        except Exception as e:  # noqa: BLE001
            try:
                _fs.delete(file_id)
            except Exception:
                pass
            _log.exception("upload: python parser crashed")
            raise HTTPException(500, f"Parser error — {e.__class__.__name__}: {str(e)[:200]}")
    else:
        # Explicit AI path — much more expensive but robust to weird PDFs.
        try:
            t_extract = time.monotonic()
            text = _extract_text(blob, ct, fname)
            _log.info("upload: extracted %d chars in %.1fs", len(text), time.monotonic() - t_extract)
        except HTTPException:
            try:
                _fs.delete(file_id)
            except Exception:
                pass
            raise
        except Exception as e:  # noqa: BLE001
            try:
                _fs.delete(file_id)
            except Exception:
                pass
            _log.exception("upload: text extraction failed")
            raise HTTPException(400, f"Could not read the file — {e.__class__.__name__}: {str(e)[:200]}")

        if not text or not text.strip():
            try:
                _fs.delete(file_id)
            except Exception:
                pass
            raise HTTPException(
                400,
                "The file contained no readable text. Scanned/image-only PDFs "
                "aren't supported yet — please export a text-based PDF or CSV "
                "from your online banking.",
            )

        try:
            parsed = await _classify(text, location_id)
            strategy = "ai"
            _enrich_with_invoices(parsed["transactions"], location_id)
        except HTTPException as he:
            try:
                _fs.delete(file_id)
            except Exception:
                pass
            _log.warning("upload: classify failed after %.1fs · %s", time.monotonic() - upload_t0, he.detail)
            raise
        except Exception as e:  # noqa: BLE001
            try:
                _fs.delete(file_id)
            except Exception:
                pass
            _log.exception("upload: unexpected classify error")
            raise HTTPException(500, f"Unexpected AI error: {e.__class__.__name__}: {str(e)[:200]}")

    txns = parsed["transactions"]
    total_income = round(sum(t["amount"] for t in txns if t["type"] == "income"), 2)
    total_expense = round(sum(t["amount"] for t in txns if t["type"] == "expense"), 2)

    doc = {
        "id": str(uuid.uuid4())[:12],
        "location_id": location_id,
        "filename": fname,
        "content_type": ct,
        "size": len(blob),
        "file_id": file_id,
        "period_start": parsed["period_start"],
        "period_end": parsed["period_end"],
        "account_ref": parsed["account_ref"],
        "currency": parsed["currency"],
        "transactions": txns,
        "income_count": sum(1 for t in txns if t["type"] == "income"),
        "expense_count": sum(1 for t in txns if t["type"] == "expense"),
        "total_income": total_income,
        "total_expense": total_expense,
        "net": round(total_income - total_expense, 2),
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "uploaded_by": user.get("email", ""),
        "uploaded_by_name": user.get("name", ""),
        # engine records how this statement was classified so the UI can
        # decide whether to offer the "Re-classify with AI" upgrade path.
        "engine": engine,       # 'python' or 'ai'
        "strategy": strategy,   # 'pypdf' / 'pdfplumber' / 'csv' / 'xlsx' / 'ai'
    }
    statements.insert_one(dict(doc))
    _log.info(
        "upload: OK (%s / %s) in %.1fs · %d income (£%.2f) · %d expense (£%.2f) · net £%.2f",
        engine, strategy,
        time.monotonic() - upload_t0,
        doc["income_count"], doc["total_income"],
        doc["expense_count"], doc["total_expense"], doc["net"],
    )
    return _strip(doc)


@router.get("/{sid}")
async def get_statement(sid: str, user: dict = Depends(get_admin_user)):
    rec = statements.find_one({"id": sid})
    if not rec:
        raise HTTPException(404, "Not found")
    return _strip(rec)


@router.post("/{sid}/reclassify")
async def reclassify_statement(
    sid: str,
    engine: str = "python",  # 'python' (default) | 'ai'
    user: dict = Depends(get_admin_user),
):
    """Re-run the classifier on a statement's stored file — useful when
    the known-supplier list has grown, category rules have improved, or
    the user wants to escalate a Python-parsed statement to the AI
    classifier for a second opinion.
    """
    rec = statements.find_one({"id": sid})
    if not rec:
        raise HTTPException(404, "Not found")

    fid = rec.get("file_id")
    if not fid:
        raise HTTPException(
            400,
            "Original file no longer available for this statement — "
            "please delete and re-upload instead.",
        )

    try:
        gf = _fs.get(fid if isinstance(fid, ObjectId) else ObjectId(str(fid)))
        blob = gf.read()
    except Exception as e:
        raise HTTPException(500, f"Could not read stored file: {e.__class__.__name__}")

    ct = rec.get("content_type") or "application/octet-stream"
    fname = rec.get("filename") or "statement"
    location_id = rec.get("location_id") or ""
    if not location_id:
        raise HTTPException(400, "Statement has no location_id — cannot re-classify")

    engine = (engine or "python").lower()
    if engine not in ("python", "ai"):
        engine = "python"
    _log.info("reclassify: sid=%s file=%r size=%d loc=%s engine=%s",
              sid, fname, len(blob), location_id, engine)

    t0 = time.monotonic()
    strategy = ""
    if engine == "python":
        try:
            suppliers = _known_suppliers(location_id)
            custom_rules = _load_custom_rules()
            parsed = bank_statement_parser.parse_locally(
                blob=blob, content_type=ct, filename=fname,
                suppliers=suppliers, supplier_matcher=_match_supplier,
                custom_rules=custom_rules,
            )
            strategy = parsed.pop("_strategy", "python")
        except ValueError as ve:
            raise HTTPException(422, str(ve))
    else:
        try:
            text = _extract_text(blob, ct, fname)
        except HTTPException:
            raise
        except Exception as e:  # noqa: BLE001
            raise HTTPException(400, f"Could not re-read the file — {e.__class__.__name__}: {str(e)[:200]}")
        if not text or not text.strip():
            raise HTTPException(400, "The stored file contained no readable text")
        parsed = await _classify(text, location_id)
        strategy = "ai"

    txns = parsed["transactions"]
    # Match each expense against real invoices for this location. Both
    # python and AI paths funnel through here so the "Matched Invoice"
    # column stays consistent regardless of engine.
    _enrich_with_invoices(txns, location_id)
    total_income = round(sum(t["amount"] for t in txns if t["type"] == "income"), 2)
    total_expense = round(sum(t["amount"] for t in txns if t["type"] == "expense"), 2)

    updated_fields = {
        "period_start": parsed["period_start"],
        "period_end": parsed["period_end"],
        "account_ref": parsed["account_ref"],
        "currency": parsed["currency"],
        "transactions": txns,
        "income_count": sum(1 for t in txns if t["type"] == "income"),
        "expense_count": sum(1 for t in txns if t["type"] == "expense"),
        "total_income": total_income,
        "total_expense": total_expense,
        "net": round(total_income - total_expense, 2),
        "engine": engine,
        "strategy": strategy,
        "reclassified_at": datetime.now(timezone.utc).isoformat(),
        "reclassified_by": user.get("email", ""),
        "reclassified_by_name": user.get("name", ""),
    }
    statements.update_one({"id": sid}, {"$set": updated_fields})

    _log.info(
        "reclassify: OK (%s/%s) in %.1fs · %d income (£%.2f) · %d expense (£%.2f)",
        engine, strategy,
        time.monotonic() - t0,
        updated_fields["income_count"], total_income,
        updated_fields["expense_count"], total_expense,
    )

    rec2 = statements.find_one({"id": sid})
    return _strip(rec2)


def _build_split_xlsx(income_txns: list, expense_txns: list, summary_rows: list) -> io.BytesIO:
    """Shared XLSX builder — single or aggregated statements land here.

    Produces the canonical 3-tab workbook: Income (green header),
    Expenses (red header) and Summary (metadata). Each transaction sheet
    has the Statement + Site columns appended when data is aggregated
    across multiple statements (i.e. any row carries a `statement_id`).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    # openpyxl refuses ASCII control characters (except \t, \n, \r) with
    # IllegalCharacterError. PDF text extraction routinely injects
    # things like \x00, \x0c (form-feed between pages), \x1f, etc. into
    # the description field, which historically caused the entire XLSX
    # download to 500 in prod. Strip them here so cells always accept
    # the value. Also normalise datetimes to ISO strings.
    _ILLEGAL_XL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

    def _s(v):
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.isoformat()
        if not isinstance(v, str):
            return v  # numbers / bools pass through
        return _ILLEGAL_XL.sub("", v)

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    income_fill = PatternFill("solid", fgColor="34C759")
    expense_fill = PatternFill("solid", fgColor="FF3B30")
    center = Alignment(horizontal="center")

    aggregated = any("statement_id" in (t or {}) for t in income_txns + expense_txns)

    def _write_sheet(ws, title, txns, fill):
        ws.title = title
        # Expenses tab gets an extra "Rule fired" column so managers can
        # see which Python regex matched each row (or "no match → other"
        # when the description didn't match any rule — a big red flag
        # that a new rule is needed). Expenses also show "Matched
        # Invoice" (a link back to the paperwork) instead of just the
        # supplier name.
        is_expenses = title == "Expenses"
        supplier_col_label = "Matched Invoice" if is_expenses else "Matched Supplier"
        headers = ["Date", "Description", "Category"]
        if is_expenses:
            headers.append("Rule fired")
        headers += [supplier_col_label, "Amount"]
        if aggregated:
            headers += ["Statement", "Site"]
        widths = [14, 56, 18] + ([32] if is_expenses else []) + [28, 14] + ([28, 22] if aggregated else [])
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = center
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        for r, t in enumerate(txns, start=2):
            ws.cell(row=r, column=1, value=_s(t.get("date", "")))
            ws.cell(row=r, column=2, value=_s(t.get("description", "")))
            ws.cell(row=r, column=3, value=_s(t.get("category", "")))
            col = 4
            if is_expenses:
                ws.cell(row=r, column=col, value=_s(t.get("category_rule", "—")))
                col += 1
            # Expenses show the matched invoice ref (Supplier #INVNUM) so
            # managers can trace which paperwork corresponds to each row.
            # Only render the ref when there IS a linked invoice id —
            # otherwise the supplier fuzzy-match sits in the Supplier
            # pivot but must NOT be presented as a real invoice link.
            if is_expenses:
                ws.cell(
                    row=r, column=col,
                    value=_s(t.get("matched_invoice_ref", "") if t.get("matched_invoice_id") else ""),
                )
            else:
                ws.cell(row=r, column=col, value=_s(t.get("matched_supplier", "")))
            col += 1
            amt_cell = ws.cell(row=r, column=col, value=float(t.get("amount") or 0))
            amt_cell.number_format = '"£"#,##0.00'
            col += 1
            if aggregated:
                ws.cell(row=r, column=col, value=_s(t.get("statement_filename", "")))
                col += 1
                ws.cell(row=r, column=col, value=_s(t.get("location_id", "")))
        if txns:
            amt_col = 6 if is_expenses else 5
            total_row = len(txns) + 2
            ws.cell(row=total_row, column=amt_col - 1, value="Total").font = Font(bold=True)
            total_cell = ws.cell(row=total_row, column=amt_col, value=sum(float(t.get("amount") or 0) for t in txns))
            total_cell.font = Font(bold=True)
            total_cell.number_format = '"£"#,##0.00'
        ws.freeze_panes = "A2"

    ws1 = wb.active
    _write_sheet(ws1, "Income", income_txns, income_fill)
    ws2 = wb.create_sheet(title="Expenses")
    _write_sheet(ws2, "Expenses", expense_txns, expense_fill)

    # ---- Expenses by Category & Supplier — pivot summaries ----
    total_exp = round(sum(float(t.get("amount") or 0) for t in expense_txns), 2)

    def _group_expenses(key_fn, title, unmatched_label, include_rules=False):
        """Group expense txns by a key function, sort by spend DESC, write a
        pivot sheet with Category/Supplier · Count · Total · Share (%).

        If `include_rules` is True (Category pivot), an extra 'Rules fired'
        column joins the distinct `category_rule` labels that contributed
        to each category — an at-a-glance audit of which regexes fired
        and where the 'no match → other' bucket is bleeding from.
        """
        agg: dict = {}
        for t in expense_txns:
            k = (key_fn(t) or "").strip() or unmatched_label
            row = agg.setdefault(k, {"count": 0, "total": 0.0, "rules": {}})
            row["count"] += 1
            row["total"] += float(t.get("amount") or 0)
            if include_rules:
                rule = (t.get("category_rule") or "no match → other").strip()
                row["rules"][rule] = row["rules"].get(rule, 0) + 1
        rows_sorted = sorted(agg.items(), key=lambda kv: kv[1]["total"], reverse=True)

        ws = wb.create_sheet(title=title)
        headers = [title.split(" by ")[-1], "Transactions", "Total", "Share"]
        widths = [30, 14, 16, 12]
        if include_rules:
            headers.append("Rules fired (count)")
            widths.append(48)
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = expense_fill
            cell.alignment = center
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        for r, (k, v) in enumerate(rows_sorted, start=2):
            ws.cell(row=r, column=1, value=_s(k))
            ws.cell(row=r, column=2, value=int(v["count"]))
            tc = ws.cell(row=r, column=3, value=round(v["total"], 2))
            tc.number_format = '"£"#,##0.00'
            share = (v["total"] / total_exp) if total_exp else 0
            sc = ws.cell(row=r, column=4, value=share)
            sc.number_format = "0.0%"
            if include_rules:
                rules = sorted(v["rules"].items(), key=lambda x: x[1], reverse=True)
                ws.cell(row=r, column=5, value=_s(", ".join(f"{name} ({cnt})" for name, cnt in rules)))

        if rows_sorted:
            total_row = len(rows_sorted) + 2
            ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
            ws.cell(row=total_row, column=2, value=sum(v["count"] for _, v in rows_sorted)).font = Font(bold=True)
            gt = ws.cell(row=total_row, column=3, value=round(total_exp, 2))
            gt.font = Font(bold=True)
            gt.number_format = '"£"#,##0.00'
            sc = ws.cell(row=total_row, column=4, value=1 if total_exp else 0)
            sc.font = Font(bold=True)
            sc.number_format = "0.0%"
        ws.freeze_panes = "A2"

    _group_expenses(lambda t: t.get("category"), "Expenses by Category", "uncategorised", include_rules=True)
    _group_expenses(lambda t: t.get("matched_supplier"), "Expenses by Supplier", "Unmatched")

    ws3 = wb.create_sheet(title="Summary")
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 34
    for r, (label, val) in enumerate(summary_rows, start=1):
        ws3.cell(row=r, column=1, value=_s(label)).font = Font(bold=True)
        ws3.cell(row=r, column=2, value=_s(val))
        if any(key in label.lower() for key in ("total", "net")):
            ws3.cell(row=r, column=2).number_format = '"£"#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/{sid}/xlsx")
async def download_xlsx(sid: str, user: dict = Depends(get_admin_user)):
    rec = statements.find_one({"id": sid})
    if not rec:
        raise HTTPException(404, "Not found")

    try:
        income_txns = [t for t in rec.get("transactions", []) if t.get("type") == "income"]
        expense_txns = [t for t in rec.get("transactions", []) if t.get("type") == "expense"]
        summary_rows = [
            ("File", rec.get("filename", "")),
            ("Location", rec.get("location_id", "")),
            ("Period start", rec.get("period_start", "")),
            ("Period end", rec.get("period_end", "")),
            ("Account", rec.get("account_ref", "")),
            ("Currency", rec.get("currency", "GBP")),
            ("Income transactions", rec.get("income_count", 0)),
            ("Expense transactions", rec.get("expense_count", 0)),
            ("Total income", float(rec.get("total_income") or 0)),
            ("Total expense", float(rec.get("total_expense") or 0)),
            ("Net", float(rec.get("net") or 0)),
            ("Uploaded at", rec.get("uploaded_at", "")),
            ("Uploaded by", rec.get("uploaded_by_name") or rec.get("uploaded_by", "")),
        ]
        buf = _build_split_xlsx(income_txns, expense_txns, summary_rows)
    except Exception as ex:  # pragma: no cover — surfaced to caller
        _log.exception("XLSX build failed for statement sid=%s file=%r", sid, rec.get("filename"))
        raise HTTPException(500, f"XLSX build failed: {type(ex).__name__}: {ex}") from ex

    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", (rec.get("filename") or "statement").rsplit(".", 1)[0])[:60]
    out_name = f"{stem}_split.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{out_name}"'},
    )


@router.get("/{sid}/file")
async def download_original(sid: str, user: dict = Depends(get_admin_user)):
    rec = statements.find_one({"id": sid})
    if not rec:
        raise HTTPException(404, "Not found")
    fid = rec.get("file_id")
    if not fid:
        raise HTTPException(404, "File missing")
    try:
        gf = _fs.get(fid if isinstance(fid, ObjectId) else ObjectId(str(fid)))
    except Exception:
        raise HTTPException(404, "File missing")
    return StreamingResponse(
        io.BytesIO(gf.read()),
        media_type=rec.get("content_type") or "application/octet-stream",
        headers={"Content-Disposition": f'inline; filename="{rec.get("filename") or "statement"}"'},
    )


@router.delete("/{sid}")
async def delete_statement(sid: str, user: dict = Depends(get_admin_user)):
    rec = statements.find_one({"id": sid})
    if not rec:
        raise HTTPException(404, "Not found")
    fid = rec.get("file_id")
    if fid:
        try:
            _fs.delete(fid if isinstance(fid, ObjectId) else ObjectId(str(fid)))
        except Exception:
            pass
    statements.delete_one({"id": sid})
    return {"deleted": True}
