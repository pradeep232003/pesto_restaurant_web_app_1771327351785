"""
Business Intelligence (BI) — admin + super_admin.

Aggregates Daily Sales + Staff Hours + Staff hourly_rate + Menu Recipes
into KPI metrics (Labour %, Food Cost %, Revenue) per location and overall.

Also exposes `/ai-insights` which sends the rollup to Claude Sonnet 4.5 and
returns a structured analysis (headline, strengths, risks, actions, anomalies).
Results are cached for 30 minutes per (period, location_id) to avoid burning
LLM credits on every page refresh.

Designed to be pure read-only — no mutations.
"""
import hashlib
import io
import json
import os
import re
import uuid
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse

from db import db, daily_sales_collection, menu_items_collection, locations_collection
from auth import get_admin_user

invoices_collection = db["invoices"]

router = APIRouter(prefix="/api/admin/bi", tags=["bi"])

staff_collection = db["staff_members"]
ai_cache = db["bi_ai_insights_cache"]
AI_CACHE_TTL = timedelta(minutes=30)


def _hours_between(start: str, end: str) -> float:
    if not start or not end:
        return 0.0
    try:
        st = datetime.strptime(start, "%H:%M")
        et = datetime.strptime(end, "%H:%M")
        diff = (et - st).total_seconds() / 3600.0
        # Handle overnight shifts (end < start)
        if diff < 0:
            diff += 24
        return round(diff, 2)
    except ValueError:
        return 0.0


def _build_rate_lookup() -> dict:
    """Build lowercase-name → hourly_rate map from staff_members collection."""
    rates = {}
    for s in staff_collection.find({}, {"_id": 0, "name": 1, "hourly_rate": 1}):
        nm = (s.get("name") or "").strip().lower()
        if nm:
            rates[nm] = float(s.get("hourly_rate") or 0)
    return rates


def _avg_recipe_cost_by_location() -> dict:
    """
    Compute the average recipe cost per menu item, grouped by location.
    Returns: {location_id: {"avg_cost": x, "avg_price": y, "items_with_recipe": n, "total_items": m}}
    Only includes items with at least one recipe line that has cost > 0.
    """
    result = {}
    cursor = menu_items_collection.find(
        {},
        {"_id": 0, "location_id": 1, "price": 1, "recipe": 1, "is_available": 1},
    )
    for it in cursor:
        loc = it.get("location_id") or "unknown"
        bucket = result.setdefault(loc, {
            "items_with_recipe": 0,
            "total_items": 0,
            "sum_cost": 0.0,
            "sum_price": 0.0,
        })
        bucket["total_items"] += 1
        recipe = it.get("recipe") or []
        cost = 0.0
        for line in recipe:
            qty = float(line.get("qty") or 0)
            uc = float(line.get("unit_cost") or 0)
            cost += qty * uc
        if cost > 0:
            bucket["items_with_recipe"] += 1
            bucket["sum_cost"] += cost
            bucket["sum_price"] += float(it.get("price") or 0)

    out = {}
    for loc, b in result.items():
        n = b["items_with_recipe"]
        out[loc] = {
            "avg_cost": round(b["sum_cost"] / n, 2) if n else 0.0,
            "avg_price": round(b["sum_price"] / n, 2) if n else 0.0,
            "items_with_recipe": n,
            "total_items": b["total_items"],
            # avg recipe-cost-to-price ratio (food cost %), only over items with recipes
            "avg_food_cost_pct": round((b["sum_cost"] / b["sum_price"]) * 100, 1) if b["sum_price"] > 0 else 0.0,
        }
    return out


def _stock_spend_by_location(start_date: str, end_date: str, location_id: Optional[str]) -> dict:
    """Sum invoice totals for category=='stock' in the given date range,
    grouped by location_id.

    Date-matching rule (matters when an invoice from March is scanned in
    June): prefer the printed invoice_date when present; only fall back
    to uploaded_at for scans that have NO printed invoice_date. This
    keeps backdated scans out of unrelated reporting periods."""
    end_upload = end_date + "T23:59:59.999Z"
    in_window = {"$gte": start_date, "$lte": end_date}
    missing_inv_date = {"$or": [{"invoice_date": ""}, {"invoice_date": {"$exists": False}}, {"invoice_date": None}]}
    q: dict = {
        "category": "stock",
        "$or": [
            {"invoice_date": in_window},
            {"$and": [missing_inv_date, {"uploaded_at": {"$gte": start_date, "$lte": end_upload}}]},
        ],
    }
    if location_id:
        q["location_id"] = location_id
    out: dict = {}
    for r in invoices_collection.find(q, {"_id": 0, "location_id": 1, "total": 1}):
        loc = r.get("location_id") or "unknown"
        out[loc] = out.get(loc, 0.0) + float(r.get("total") or 0.0)
    return out



@router.get("")
async def bi_overview(
    start_date: Optional[str] = Query(None, description="YYYY-MM-DD inclusive"),
    end_date: Optional[str] = Query(None, description="YYYY-MM-DD inclusive"),
    location_id: Optional[str] = Query(None),
    user: dict = Depends(get_admin_user),
):
    """
    Returns BI overview:
    - kpi: { total_revenue, total_labour, labour_pct, est_food_cost_pct, total_hours }
    - by_location: list of per-location rollups
    - menu: avg food cost % per location (recipe-based)
    - period: { start_date, end_date, days }
    """
    return _compute_overview(start_date, end_date, location_id)


def _compute_overview(start_date: Optional[str], end_date: Optional[str], location_id: Optional[str]) -> dict:
    """Internal: build the BI overview dict so other endpoints (AI insights)
    can reuse it without going through HTTP."""
    # Default to last 7 days inclusive
    if not start_date or not end_date:
        today = datetime.now(timezone.utc).date()
        end_date = end_date or today.isoformat()
        start_date = start_date or (today - timedelta(days=6)).isoformat()

    sales_query: dict = {"date": {"$gte": start_date, "$lte": end_date}}
    if location_id:
        sales_query["location_id"] = location_id

    entries = list(daily_sales_collection.find(sales_query, {"_id": 0}))
    rates = _build_rate_lookup()
    menu_rollup = _avg_recipe_cost_by_location()
    # Per-location stock spend from invoices in the same window — the true
    # food cost ground truth (anything tagged "stock"). Falls back to the
    # recipe-based estimate per-location if a site has no scans yet.
    stock_spend_by_loc = _stock_spend_by_location(start_date, end_date, location_id)

    # Per-location accumulators
    per_loc: dict = {}
    total_revenue = 0.0
    total_labour = 0.0
    total_hours = 0.0

    for e in entries:
        loc = e.get("location_id") or "unknown"
        bucket = per_loc.setdefault(loc, {
            "location_id": loc,
            "revenue": 0.0,
            "labour": 0.0,
            "hours": 0.0,
            "days": 0,
            "staff_breakdown": {},  # name -> {hours, cost}
        })

        rev = float(e.get("sales") or 0)
        bucket["revenue"] += rev
        bucket["days"] += 1
        total_revenue += rev

        for sh in (e.get("staff_hours") or []):
            name = (sh.get("name") or "").strip()
            if not name:
                continue
            hrs = _hours_between(sh.get("start_time", ""), sh.get("end_time", ""))
            rate = rates.get(name.lower(), 0.0)
            cost = round(hrs * rate, 2)
            bucket["hours"] += hrs
            bucket["labour"] += cost
            sb = bucket["staff_breakdown"].setdefault(name, {"hours": 0.0, "cost": 0.0, "rate": rate})
            sb["hours"] = round(sb["hours"] + hrs, 2)
            sb["cost"] = round(sb["cost"] + cost, 2)
            total_hours += hrs
            total_labour += cost

    # Finalise per-location records
    by_location = []
    loc_names = {ldoc["id"]: ldoc.get("name", ldoc["id"]) for ldoc in locations_collection.find({}, {"_id": 0, "id": 1, "name": 1})}
    for loc, b in per_loc.items():
        rev = b["revenue"]
        lab = b["labour"]
        labour_pct = round((lab / rev) * 100, 1) if rev > 0 else 0.0
        menu = menu_rollup.get(loc, {"avg_food_cost_pct": 0.0, "items_with_recipe": 0, "total_items": 0})
        # Prefer real invoice spend; fall back to recipe estimate when no
        # invoices have been scanned for this site in the window. This
        # makes "Food %" reflect what the manager actually spent, not a
        # theoretical menu ratio.
        invoice_food_cost = round(stock_spend_by_loc.get(loc, 0.0), 2)
        if invoice_food_cost > 0:
            est_food_cost = invoice_food_cost
            food_cost_pct = round((est_food_cost / rev) * 100, 1) if rev > 0 else 0.0
            food_cost_source = "invoices"
        else:
            est_food_cost = round((rev * menu["avg_food_cost_pct"] / 100), 2) if menu["avg_food_cost_pct"] > 0 else 0.0
            food_cost_pct = menu["avg_food_cost_pct"]
            food_cost_source = "recipes"
        gross_margin = round(rev - lab - est_food_cost, 2)
        by_location.append({
            "location_id": loc,
            "location_name": loc_names.get(loc, loc),
            "revenue": round(rev, 2),
            "labour": round(lab, 2),
            "labour_pct": labour_pct,
            "hours": round(b["hours"], 2),
            "days": b["days"],
            "food_cost_pct": food_cost_pct,
            "est_food_cost": est_food_cost,
            "food_cost_source": food_cost_source,
            "invoice_stock_spend": invoice_food_cost,
            "gross_margin": gross_margin,
            "gross_margin_pct": round((gross_margin / rev) * 100, 1) if rev > 0 else 0.0,
            "menu_coverage": {
                "items_with_recipe": menu["items_with_recipe"],
                "total_items": menu["total_items"],
            },
            "staff_breakdown": sorted(
                [{"name": n, **v} for n, v in b["staff_breakdown"].items()],
                key=lambda x: x["cost"], reverse=True,
            ),
        })
    by_location.sort(key=lambda x: x["revenue"], reverse=True)

    # Overall food cost % — sum of per-location est_food_cost / total_revenue.
    # The per-location est_food_cost already prefers invoice spend where
    # available; the overall source label is "invoices" iff every location
    # used invoices (mixed → "mixed").
    total_food_cost = sum(loc["est_food_cost"] for loc in by_location)
    overall_food_cost_pct = round((total_food_cost / total_revenue) * 100, 1) if total_revenue > 0 else 0.0
    overall_labour_pct = round((total_labour / total_revenue) * 100, 1) if total_revenue > 0 else 0.0
    overall_gross = round(total_revenue - total_labour - total_food_cost, 2)
    sources = {loc["food_cost_source"] for loc in by_location if loc["est_food_cost"] > 0}
    if sources == {"invoices"}:
        overall_food_source = "invoices"
    elif sources == {"recipes"}:
        overall_food_source = "recipes"
    elif sources:
        overall_food_source = "mixed"
    else:
        overall_food_source = "none"
    total_invoice_stock_spend = sum(loc.get("invoice_stock_spend", 0.0) for loc in by_location)

    # Period length in days
    try:
        sd = datetime.strptime(start_date, "%Y-%m-%d").date()
        ed = datetime.strptime(end_date, "%Y-%m-%d").date()
        days = (ed - sd).days + 1
    except ValueError:
        days = len({e.get("date") for e in entries})

    return {
        "period": {"start_date": start_date, "end_date": end_date, "days": days},
        "kpi": {
            "total_revenue": round(total_revenue, 2),
            "total_labour": round(total_labour, 2),
            "total_hours": round(total_hours, 2),
            "labour_pct": overall_labour_pct,
            "est_food_cost": round(total_food_cost, 2),
            "food_cost_pct": overall_food_cost_pct,
            "food_cost_source": overall_food_source,
            "invoice_stock_spend": round(total_invoice_stock_spend, 2),
            "gross_margin": overall_gross,
            "gross_margin_pct": round((overall_gross / total_revenue) * 100, 1) if total_revenue > 0 else 0.0,
            "entries": len(entries),
        },
        "by_location": by_location,
        "menu_recipe_summary": menu_rollup,
    }


@router.get("/menu-cost")
async def menu_cost_breakdown(
    location_id: Optional[str] = Query(None),
    user: dict = Depends(get_admin_user),
):
    """
    Per-menu-item recipe cost & margin breakdown — admin+.
    Returns list of items with computed cost, margin £, margin %.
    """
    query: dict = {}
    if location_id:
        query["location_id"] = location_id

    items_out = []
    for it in menu_items_collection.find(query, {"_id": 0}):
        recipe = it.get("recipe") or []
        cost = 0.0
        for line in recipe:
            qty = float(line.get("qty") or 0)
            uc = float(line.get("unit_cost") or 0)
            cost += qty * uc
        price = float(it.get("price") or 0)
        margin = price - cost
        items_out.append({
            "id": it.get("id"),
            "name": it.get("name"),
            "location_id": it.get("location_id"),
            "category": it.get("category"),
            "price": round(price, 2),
            "recipe_cost": round(cost, 2),
            "margin": round(margin, 2),
            "food_cost_pct": round((cost / price) * 100, 1) if price > 0 else 0.0,
            "margin_pct": round((margin / price) * 100, 1) if price > 0 else 0.0,
            "recipe_lines": len(recipe),
            "has_recipe": cost > 0,
        })

    items_out.sort(key=lambda x: (not x["has_recipe"], -x["margin"]))
    return {"items": items_out}



# ---------------------------------------------------------------------------
# AI Insights — Claude Sonnet 4.5 analysis of the BI rollup.
# ---------------------------------------------------------------------------

_AI_SYSTEM_PROMPT = """You are a senior restaurant business analyst for a multi-site
quick-service food brand. You are given an aggregated BI dataset (revenue,
labour cost, labour %, estimated food cost, gross margin, hours per site, top
staff by cost) for a specific date window.

Industry benchmarks for QSR/cafes:
- Healthy labour cost: 25-30% of revenue (>35% = critical)
- Healthy food cost: 28-32% of revenue (>35% = critical)
- Healthy gross margin (revenue - labour - food): 35%+ (negative = critical)

Your job: return a JSON object with concise, actionable insights. NEVER invent
numbers — only reason from the data provided. Use British English (e.g. "labour"
not "labor", "£" for currency). Keep each bullet under 20 words.

Return ONLY valid JSON in this exact shape (no markdown, no prose):
{
  "headline": "One punchy sentence summarising overall business health",
  "health_score": 0-100,
  "health_label": "Excellent" | "Strong" | "Healthy" | "At risk" | "Critical",
  "strengths": ["...", "..."],
  "risks": ["...", "..."],
  "actions": [
    {"priority": "high" | "medium" | "low", "title": "...", "detail": "...", "impact": "..."}
  ],
  "anomalies": ["Specific outliers or surprising patterns, if any"]
}
"""


def _cache_key(start_date: str, end_date: str, location_id: Optional[str], overview_digest: str) -> str:
    base = f"{start_date}|{end_date}|{location_id or 'all'}|{overview_digest}"
    return hashlib.sha256(base.encode("utf-8")).hexdigest()[:24]


def _summarise_for_llm(overview: dict) -> dict:
    """Trim the overview to the fields Claude actually needs. Keeps the prompt
    small (cost) and the model focused (quality)."""
    locs = []
    for loc in overview.get("by_location", []):
        # Keep only top 3 staff per location by cost — anything below that is
        # noise for a high-level analysis.
        top_staff = sorted(loc.get("staff_breakdown", []), key=lambda s: -s.get("cost", 0))[:3]
        locs.append({
            "name": loc.get("location_name"),
            "revenue": loc.get("revenue"),
            "labour": loc.get("labour"),
            "labour_pct": loc.get("labour_pct"),
            "est_food_cost": loc.get("est_food_cost"),
            "food_cost_pct": loc.get("food_cost_pct"),
            "gross_margin": loc.get("gross_margin"),
            "gross_margin_pct": loc.get("gross_margin_pct"),
            "hours": loc.get("hours"),
            "days_traded": loc.get("days"),
            "menu_recipe_coverage": loc.get("menu_coverage"),
            "top_staff_cost": [{"name": s["name"], "hours": s["hours"], "cost": s["cost"]} for s in top_staff],
        })
    return {
        "period": overview.get("period"),
        "kpi": overview.get("kpi"),
        "locations": locs,
    }


def _extract_json(text: str) -> dict:
    """Tolerate model output wrapped in ```json fences or with leading prose
    by hunting for the first/last brace pair before parsing."""
    s = text.strip()
    # Strip code fences
    if s.startswith("```"):
        s = re.sub(r"^```(?:json)?\s*", "", s)
        s = re.sub(r"\s*```$", "", s)
    # If still not pure JSON, find the outermost braces.
    if not s.lstrip().startswith("{"):
        m = re.search(r"\{.*\}", s, re.DOTALL)
        if m:
            s = m.group(0)
    return json.loads(s)


async def _generate_insights(overview: dict) -> dict:
    """Call Claude Sonnet 4.5 directly via Anthropic's REST API using httpx.

    We deliberately avoid any external SDK so the function works on any host
    (Railway, Vercel, bare VPS) without installing extra packages — httpx is
    already a FastAPI transitive dependency."""
    from routes.ai_settings import get_active_ai_key, get_active_ai_provider
    api_key = get_active_ai_key()
    if not api_key:
        raise HTTPException(
            500,
            "AI insights unavailable: no API key configured. Open Admin → AI Settings to add one.",
        )
    provider = get_active_ai_provider()

    import httpx

    payload = _summarise_for_llm(overview)
    user_text = (
        "Analyse the following Jolly's Kafe BI data window and return your JSON. "
        "Focus on: where money is being made/lost, labour vs food-cost imbalance, "
        "underperforming sites, and the 3-5 highest-impact actions a manager can "
        "take this week.\n\n"
        f"DATA:\n{json.dumps(payload, indent=2)}"
    )

    req = {
        "model": "claude-sonnet-4-5-20250929",
        "max_tokens": 2048,
        "system": _AI_SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": user_text}],
    }
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }

    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post("https://api.anthropic.com/v1/messages", json=req, headers=headers)
    except HTTPException:
        raise
    except Exception as e:  # noqa: BLE001
        import logging
        logging.exception("BI AI insights HTTP call failed")
        snippet = (str(e) or e.__class__.__name__).splitlines()[0][:240]
        raise HTTPException(502, f"AI provider error ({provider}): {snippet}")

    if resp.status_code >= 400:
        # Surface Anthropic's own error verbatim — much more useful than a generic message.
        try:
            err = resp.json().get("error", {}).get("message", resp.text)
        except Exception:
            err = resp.text
        raise HTTPException(502, f"AI provider error ({provider}, {resp.status_code}): {err[:240]}")

    data = resp.json()
    full = "".join(
        block.get("text", "") for block in (data.get("content") or [])
        if block.get("type") == "text"
    ).strip()
    if not full:
        raise HTTPException(502, "AI provider returned an empty response")

    try:
        return _extract_json(full)
    except json.JSONDecodeError as e:
        raise HTTPException(502, f"AI returned non-JSON response: {e}")


@router.get("/ai-insights")
async def bi_ai_insights(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    location_id: Optional[str] = Query(None),
    refresh: bool = Query(False, description="Bypass the 30-min cache"),
    user: dict = Depends(get_admin_user),
):
    overview = _compute_overview(start_date, end_date, location_id)

    # If there's literally no revenue in the window, skip the LLM call and
    # return a deterministic "no data" envelope so the page renders sanely.
    if not overview["kpi"]["total_revenue"]:
        return {
            "period": overview["period"],
            "insights": {
                "headline": "No sales data in this window.",
                "health_score": 0,
                "health_label": "No data",
                "strengths": [],
                "risks": ["No daily sales entries logged for this period."],
                "actions": [{
                    "priority": "high",
                    "title": "Log daily sales",
                    "detail": "Staff should be filling in Daily Sales each evening — empty windows make BI useless.",
                    "impact": "Unlocks BI, labour cost, and food cost analysis for this period.",
                }],
                "anomalies": [],
            },
            "cached": False,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }

    # Cache key derives from the overview digest so any data change busts it.
    digest = hashlib.sha256(
        json.dumps(_summarise_for_llm(overview), sort_keys=True).encode("utf-8")
    ).hexdigest()[:16]
    key = _cache_key(overview["period"]["start_date"], overview["period"]["end_date"], location_id, digest)

    if not refresh:
        cached = ai_cache.find_one({"key": key}, {"_id": 0})
        if cached:
            try:
                gen_at = datetime.fromisoformat(cached["generated_at"])
                if datetime.now(timezone.utc) - gen_at < AI_CACHE_TTL:
                    return {**cached, "cached": True}
            except Exception:
                pass

    insights = await _generate_insights(overview)
    record = {
        "key": key,
        "period": overview["period"],
        "location_id": location_id,
        "insights": insights,
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }
    ai_cache.update_one({"key": key}, {"$set": record}, upsert=True)
    return {**{k: v for k, v in record.items() if k != "_id"}, "cached": False}


# ---------------------------------------------------------------------------
# Menu Engineering — Kasavana & Smith quadrant analysis.
#
# Classifies every menu item by (popularity, profitability) vs its site
# average and buckets it into Stars / Plow Horses / Puzzles / Dogs so
# owners can see at a glance what to promote, reprice, rebrand, or bin.
#
#   • Popularity  = units sold in the period, relative to the site mean.
#   • Profitability = contribution margin (price − recipe cost) per unit,
#     relative to the site mean.
#
# An item needs BOTH a recipe (to know food cost) AND at least one unit
# sold in the window to be classified — otherwise it's returned as
# `quadrant: "uncategorised"` so the manager can complete the setup.
# ---------------------------------------------------------------------------

orders_collection = db["orders"]
menu_sales_collection = db["menu_sales"]

# Any order that reached the "money in the till" stage counts as a real
# sale. Cancelled/refunded ones are ignored so they don't skew popularity.
_SALE_STATUSES = ("completed", "ready", "collected", "paid", "confirmed")


def _classify(units: float, margin: float, mean_units: float, mean_margin: float) -> str:
    high_pop = units >= mean_units
    high_prof = margin >= mean_margin
    if high_pop and high_prof:
        return "star"
    if high_pop and not high_prof:
        return "plow_horse"
    if not high_pop and high_prof:
        return "puzzle"
    return "dog"


@router.get("/menu-engineering")
async def menu_engineering(
    location_id: Optional[str] = Query(None),
    days: int = Query(30, ge=1, le=365),
    user: dict = Depends(get_admin_user),
):
    """Kasavana & Smith 2x2 quadrant analysis for one site (or all)."""
    now = datetime.now(timezone.utc)
    start = (now - timedelta(days=days)).isoformat()
    end = now.isoformat()

    # 1. Pull menu items for the location (or all), keyed by id + name.
    menu_query: dict = {}
    if location_id:
        menu_query["location_id"] = location_id
    menu_docs = list(menu_items_collection.find(
        menu_query,
        {"_id": 0, "id": 1, "name": 1, "price": 1, "recipe": 1, "location_id": 1, "is_available": 1},
    ))
    menu_by_id: dict = {}
    menu_by_name: dict = {}
    for m in menu_docs:
        recipe = m.get("recipe") or []
        cost = round(sum(float(line.get("cost") or 0) for line in recipe if isinstance(line, dict)), 2)
        price = float(m.get("price") or 0)
        menu_by_id[m.get("id")] = {
            "id": m.get("id"),
            "name": m.get("name") or "",
            "location_id": m.get("location_id"),
            "price": price,
            "cost": cost,
            "margin": round(price - cost, 2),
            "has_recipe": cost > 0,
            "is_available": bool(m.get("is_available", True)),
            "units": 0,
            "revenue": 0.0,
        }
        if m.get("name"):
            # Fallback lookup by name — order line items may lack a menu_id
            # if they came from the old EPOS import.
            menu_by_name[m["name"].strip().lower()] = menu_by_id[m.get("id")]

    # 2. Aggregate units + revenue from orders in the window.
    order_query: dict = {
        "created_at": {"$gte": start, "$lte": end},
        "status": {"$in": list(_SALE_STATUSES)},
    }
    if location_id:
        order_query["location_id"] = location_id
    for order in orders_collection.find(order_query, {"_id": 0, "items": 1}):
        for line in (order.get("items") or []):
            if not isinstance(line, dict):
                continue
            item_id = line.get("id") or line.get("menu_item_id")
            qty = float(line.get("quantity") or 0)
            price = float(line.get("price") or 0)
            record = menu_by_id.get(item_id)
            if record is None:
                # Fallback — match by name so items renamed on the menu after
                # the order was placed still count.
                nm = (line.get("name") or "").strip().lower()
                record = menu_by_name.get(nm)
            if record is None:
                continue
            record["units"] += qty
            record["revenue"] += qty * price

    # 2b. Layer in uploaded sales data (XLSX imports). Managers whose POS
    # isn't integrated can push a spreadsheet — those rows show up here
    # alongside any native orders.
    sales_query: dict = {"date": {"$gte": start[:10], "$lte": end[:10]}}
    if location_id:
        sales_query["location_id"] = location_id
    for row in menu_sales_collection.find(sales_query, {"_id": 0, "item_name": 1, "item_id": 1, "units": 1, "unit_price": 1}):
        qty = float(row.get("units") or 0)
        price = float(row.get("unit_price") or 0)
        record = None
        rid = row.get("item_id")
        if rid:
            record = menu_by_id.get(rid)
        if record is None:
            nm = (row.get("item_name") or "").strip().lower()
            record = menu_by_name.get(nm)
        if record is None:
            continue
        record["units"] += qty
        # If the upload didn't include a price, fall back to menu list price
        # so revenue still totals sensibly (unit_price will then be 0).
        record["revenue"] += qty * (price if price > 0 else record["price"])

    # 3. Compute site means over items that are actually eligible (sold AND
    # priced with a recipe). Two separate means so a dish with no recipe
    # doesn't drag the profitability mean toward zero.
    eligible = [m for m in menu_by_id.values() if m["units"] > 0 and m["has_recipe"]]
    total_units = sum(m["units"] for m in eligible)
    total_margin = sum(m["margin"] for m in eligible)   # each item counted once, per Kasavana
    unique_items = len(eligible)
    mean_units = (total_units / unique_items) if unique_items else 0.0
    mean_margin = (total_margin / unique_items) if unique_items else 0.0

    # 4. Classify every item; uncategorised = missing recipe OR zero sales.
    items_out: list = []
    counts = {"star": 0, "plow_horse": 0, "puzzle": 0, "dog": 0, "uncategorised": 0}
    for m in menu_by_id.values():
        if m["units"] <= 0 or not m["has_recipe"]:
            quadrant = "uncategorised"
            reason = ("no sales" if m["units"] <= 0 else "") + \
                     (" & " if m["units"] <= 0 and not m["has_recipe"] else "") + \
                     ("no recipe" if not m["has_recipe"] else "")
        else:
            quadrant = _classify(m["units"], m["margin"], mean_units, mean_margin)
            reason = ""
        counts[quadrant] += 1
        food_pct = round((m["cost"] / m["price"] * 100.0), 1) if m["price"] > 0 else 0.0
        items_out.append({
            **m,
            "units": round(m["units"], 2),
            "revenue": round(m["revenue"], 2),
            "food_cost_pct": food_pct,
            "quadrant": quadrant,
            "reason": reason,
        })

    # Sort: dogs first (need action), then puzzles, plow horses, stars, uncat.
    order_map = {"dog": 0, "plow_horse": 1, "puzzle": 2, "star": 3, "uncategorised": 4}
    items_out.sort(key=lambda r: (order_map.get(r["quadrant"], 9), -r["revenue"]))

    return {
        "period": {
            "start": start[:10],
            "end": end[:10],
            "days": days,
        },
        "location_id": location_id,
        "benchmarks": {
            "mean_units": round(mean_units, 2),
            "mean_margin": round(mean_margin, 2),
            "total_units": round(total_units, 2),
            "eligible_items": unique_items,
            "total_menu_items": len(menu_by_id),
        },
        "counts": counts,
        "items": items_out,
    }


# ---------------------------------------------------------------------------
# XLSX sales import — for managers whose POS isn't integrated.
# ---------------------------------------------------------------------------

_TEMPLATE_HEADERS = ["item_name", "units_sold", "unit_price", "date"]

@router.get("/menu-engineering/template")
async def menu_engineering_template(user: dict = Depends(get_admin_user)):
    """Download a blank XLSX with the exact columns we expect, plus two
    example rows so the manager knows the shape. `item_name` matches
    against the menu; `date` is optional (defaults to today when blank)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(_TEMPLATE_HEADERS)
    # Two example rows — use realistic names to hint at case-insensitive match.
    ws.append(["Chicken Katsu Curry", 12, 12.50, datetime.now(timezone.utc).strftime("%Y-%m-%d")])
    ws.append(["Vegan Buddha Bowl", 5, 11.00, datetime.now(timezone.utc).strftime("%Y-%m-%d")])
    # Widen columns a touch for readability.
    for col_letter, width in [("A", 32), ("B", 12), ("C", 12), ("D", 14)]:
        ws.column_dimensions[col_letter].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="menu-engineering-sales-template.xlsx"'},
    )


@router.post("/menu-engineering/upload")
async def menu_engineering_upload(
    file: UploadFile = File(...),
    location_id: str = Form(...),
    user: dict = Depends(get_admin_user),
):
    """Parse the uploaded XLSX and store one `menu_sales` row per line.

    Rows are grouped under a fresh `upload_id` so the admin can undo the
    entire import in one click later. Matching is by `item_name`
    (case-insensitive) against the site's menu — unmatched rows are
    returned in `unmatched` so the manager can fix the source spreadsheet.
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(415, "Please upload an XLSX file")
    blob = await file.read()
    if not blob:
        raise HTTPException(400, "Empty file")
    if len(blob) > 10 * 1024 * 1024:
        raise HTTPException(413, "File exceeds 10 MB limit")

    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(400, f"Could not read XLSX: {e}")
    ws = wb.active
    if ws is None:
        raise HTTPException(400, "Workbook has no active sheet")

    # First row is expected to be headers; be forgiving with header order.
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "File contains no data rows")
    headers = [str(c or "").strip().lower().replace(" ", "_") for c in rows[0]]
    try:
        idx_name = headers.index("item_name")
        idx_units = headers.index("units_sold")
    except ValueError:
        raise HTTPException(400, "Missing required columns: 'item_name' and 'units_sold'")
    idx_price = headers.index("unit_price") if "unit_price" in headers else -1
    idx_date = headers.index("date") if "date" in headers else -1

    # Load menu items so we can attempt an id match up front.
    menu_by_name: dict = {}
    for m in menu_items_collection.find(
        {"location_id": location_id},
        {"_id": 0, "id": 1, "name": 1},
    ):
        nm = (m.get("name") or "").strip().lower()
        if nm:
            menu_by_name[nm] = m.get("id")

    upload_id = str(uuid.uuid4())[:12]
    today_iso = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    now = datetime.now(timezone.utc).isoformat()
    to_insert: list = []
    unmatched: list = []
    matched = 0

    for r in rows[1:]:
        raw_name = (r[idx_name] if idx_name < len(r) else "") or ""
        item_name = str(raw_name).strip()
        if not item_name:
            continue
        raw_units = r[idx_units] if idx_units < len(r) else None
        try:
            units = float(raw_units or 0)
        except (TypeError, ValueError):
            units = 0.0
        if units <= 0:
            continue
        price = 0.0
        if idx_price >= 0 and idx_price < len(r):
            try:
                price = float(r[idx_price] or 0)
            except (TypeError, ValueError):
                price = 0.0
        date_iso = today_iso
        if idx_date >= 0 and idx_date < len(r):
            raw_date = r[idx_date]
            if isinstance(raw_date, datetime):
                date_iso = raw_date.strftime("%Y-%m-%d")
            elif raw_date:
                # Try a few common string formats.
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
                    try:
                        date_iso = datetime.strptime(str(raw_date).strip(), fmt).strftime("%Y-%m-%d")
                        break
                    except ValueError:
                        continue

        item_id = menu_by_name.get(item_name.lower())
        if item_id:
            matched += 1
        else:
            unmatched.append(item_name)

        to_insert.append({
            "id": str(uuid.uuid4())[:12],
            "upload_id": upload_id,
            "location_id": location_id,
            "item_name": item_name,
            "item_id": item_id,
            "units": units,
            "unit_price": price,
            "date": date_iso,
            "uploaded_at": now,
            "uploaded_by": user.get("email", ""),
        })

    if not to_insert:
        raise HTTPException(400, "No valid data rows found (need item_name + units_sold > 0)")

    menu_sales_collection.insert_many(to_insert)
    # Dedupe the unmatched list for the caller.
    unmatched_unique = sorted(set(unmatched), key=str.lower)

    return {
        "upload_id": upload_id,
        "rows_saved": len(to_insert),
        "matched": matched,
        "unmatched_count": len(unmatched_unique),
        "unmatched": unmatched_unique[:50],
    }


@router.get("/menu-engineering/uploads")
async def list_menu_engineering_uploads(
    location_id: str = Query(...),
    user: dict = Depends(get_admin_user),
):
    """List recent XLSX imports for this site so the admin can undo them."""
    pipeline = [
        {"$match": {"location_id": location_id}},
        {"$group": {
            "_id": "$upload_id",
            "rows": {"$sum": 1},
            "units": {"$sum": "$units"},
            "uploaded_at": {"$max": "$uploaded_at"},
            "uploaded_by": {"$last": "$uploaded_by"},
        }},
        {"$sort": {"uploaded_at": -1}},
        {"$limit": 30},
    ]
    docs = list(menu_sales_collection.aggregate(pipeline))
    return [{"upload_id": d["_id"], "rows": d["rows"], "units": d["units"],
             "uploaded_at": d["uploaded_at"], "uploaded_by": d["uploaded_by"]} for d in docs]


@router.delete("/menu-engineering/uploads/{upload_id}")
async def delete_menu_engineering_upload(upload_id: str, user: dict = Depends(get_admin_user)):
    r = menu_sales_collection.delete_many({"upload_id": upload_id})
    return {"deleted": r.deleted_count}
